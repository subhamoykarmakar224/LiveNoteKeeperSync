from openpyxl import *
import datetime
from Configuration import *
import mysql.connector as con

def addNote(note):
    wb = load_workbook(workbook_name)
    sheet = wb['notes']
    max = sheet.max_row
    # print(sheet.cell(row=1, column=1).value)
    sheet.cell(row=max+1, column=2).value = str(datetime.datetime.now())
    sheet.cell(row=max+1, column=3).value = note
    wb.save(filename= workbook_name)
    wb.close()


def getAllNotes():
    data = []
    wb = load_workbook(workbook_name)
    sheet = wb['notes']
    max = sheet.max_row
    for i in range(2, max+1):
        print(
            sheet.cell(row=i, column=1).value,
            sheet.cell(row=i, column=2).value,
            sheet.cell(row=i, column=3).value
        )
    wb.close()


def getDbObject():
    db = con.connect(host="localhost", user="root", passwd="cdcju", database="notekeeper", auth_plugin='mysql_native_password')
    return db


def insertDatSQL(id, note, timestamp):
    id = id['name']
    db = getDbObject()
    cur = db.cursor()
    q = "insert into " + TABLE_NOTES + " ("+KEY_NOTES_NOTEID+", "+KEY_NOTES_NOTE+", " + \
        KEY_NOTES_DATESTAMP + ", " + KEY_NOTES_HASHVAL + ", "+ KEY_NOTES_LASTCHANGED + \
        ") values('" + id + "', '" + note + "', '" + timestamp + "', '" + str(hash(timestamp)) + "', '" + timestamp + "')"
    cur.execute(q)
    db.commit()
    cur.close()
    db.close()


def getDataAsPerDate(date):
    data= {}
    db = getDbObject()
    cur = db.cursor()
    q = 'select ' + KEY_NOTES_NOTEID + ', ' + KEY_NOTES_NOTE + ', ' + KEY_NOTES_DATESTAMP + ', ' + KEY_NOTES_HASHVAL + \
        ' from notes where datestamp >= "' + date + ' 00:00:01" and datestamp <= "' + date + ' 23:59:59" order by ' + KEY_NOTES_DATESTAMP + ' desc'
    cur.execute(q)
    for c in cur:
        data[str(c[0])] = [str(c[1]), str(c[2]), str(c[3])]
    db.commit()
    cur.close()
    db.close()
    return data


def updateDataAsPerId(id, note, dm):
    db = getDbObject()
    cur = db.cursor()
    s = str(hash(str(datetime.datetime.now())))
    q = 'update ' + TABLE_NOTES + ' set note="' + note + '", ' + KEY_NOTES_HASHVAL + '="' + s + '", ' +\
        KEY_NOTES_LASTCHANGED + '="' + str(dm) + '" where ' + KEY_NOTES_NOTEID + '="' + id +'"'
    cur.execute(q)
    db.commit()
    cur.close()
    db.close()


def deleteNoteById(id, dm):
    db = getDbObject()
    cur = db.cursor()
    # q = 'delete from ' + TABLE_NOTES + ' where ' + KEY_NOTES_NOTEID + ' = "' + id + '"'
    q = 'update ' + TABLE_NOTES + ' set ' + KEY_NOTES_HASHVAL + ' = "delete", ' + KEY_NOTES_LASTCHANGED + ' ="'+ \
        dm +'" where ' + KEY_NOTES_NOTEID + '="' + id + '"'
    print(q)
    cur.execute(q)
    db.commit()
    cur.close()
    db.close()


def getAllDataCount():
    cnt = 0
    db = getDbObject()
    cur = db.cursor()
    q = 'select count(*) from ' + TABLE_NOTES
    cur.execute(q)
    for c in cur:
        cnt = int(c[0])
    db.commit()
    cur.close()
    db.close()
    return cnt


def justInsert(res):
    db = getDbObject()
    cur = db.cursor()
    for k in res.keys():
        val = res[k]
        note = str(val['note'])
        timestamp = str(val['datestamp'])
        hashVal = str(val['hashval'])
        q = "insert into " + TABLE_NOTES + " (" + KEY_NOTES_NOTEID + ", " + KEY_NOTES_NOTE + ", " + \
            KEY_NOTES_DATESTAMP + ", " + KEY_NOTES_HASHVAL + ") values('" + str(k) + "', '" + note + "', '" + timestamp + "', '" + \
            hashVal + "')"
        cur.execute(q)

    db.commit()
    db.close()


def getAllData():
    res = {}
    db = getDbObject()
    cur = db.cursor()
    q = "select " + KEY_NOTES_NOTEID + ", " + KEY_NOTES_NOTE + ", " + KEY_NOTES_DATESTAMP + ", " + KEY_NOTES_HASHVAL + ", " + KEY_NOTES_LASTCHANGED + " from " + TABLE_NOTES
    cur.execute(q)
    for c in cur:
        res[str(c[0])] = [str(c[1]), str(c[2]), str(c[3]), str(c[4])]

    cur.close()
    db.close()
    return res


def updateNoteID(oldId, newId):
    db = getDbObject()
    cur = db.cursor()
    q = 'update ' + TABLE_NOTES + ' set '+KEY_NOTES_NOTEID+'="' + newId + '" where ' + KEY_NOTES_NOTEID + '="' + oldId +'"'
    cur.execute(q)
    db.commit()
    db.close()


def finalDelete(id):
    db = getDbObject()
    cur = db.cursor()
    q = 'delete from ' + TABLE_NOTES + ' where ' + KEY_NOTES_NOTEID + '="' + id + '"'
    print(q)
    cur.execute(q)
    db.commit()
    db.close()


# def fixID():
#     db = getDbObject()
#     cur = db.cursor()
#     q = 'select ' + KEY_NOTES_NOTEID + ' from ' + TABLE_NOTES
#     cur.execute(q)
#     k = {}
#     for c in cur:
#         t = str(c[0]).split(": ")[1].strip("\'")
#         t = t[:len(t)-2]
#         k[str(c[0])] = str(t)
#
#     cur.close()
#     cur = db.cursor()
#     for i in k.keys():
#         q = 'update ' + TABLE_NOTES + ' set ' + KEY_NOTES_NOTEID + ' = "' + k[i] + '" where ' + KEY_NOTES_NOTEID + '="' + i + '"'
#         print(q)
#         cur.execute(q)
#     db.commit()
#     cur.close()
#     db.close()

# def insertLastChanged():
#     db = getDbObject()
#     cur = db.cursor()
#     q = 'select ' + KEY_NOTES_NOTEID + ', '+ KEY_NOTES_DATESTAMP +' from ' + TABLE_NOTES
#     cur.execute(q)
#     data = {}
#     for c in cur:
#         data[str(c[0])] = str(c[1])
#
#     cur.close()
#     cur = db.cursor()
#     for i in data.keys():
#         q = 'update ' + TABLE_NOTES + ' set ' + KEY_NOTES_LASTCHANGED + ' = "' + data[i] + '" where ' + KEY_NOTES_NOTEID + '="' + i + '"'
#         print(q)
#         cur.execute(q)
#     db.commit()
#     cur.close()
#     db.close()

# if __name__ == '__main__':
#     insertLastChanged()
'''
{'-M6AINfmhkzpuOwu-fjt': {'datestamp': '2020-04-30 19:06:26.773639', 'hashval': 6087446085743364701,
                          'note': 'Go to gym and do cardio'},
 '-M6AIPx8NfHwOum1C0r0': {'datestamp': '2020-04-30 19:06:36.919289', 'hashval': 1111953397585184280,
                          'note': 'Do yoga in roof'},
 '-M6AISg098mAucpmq8bO': {'datestamp': '2020-04-30 19:06:48.111315', 'hashval': -280328716692055736,
                          'note': 'Pick up newspaper'},
 '-M6AImcjT-GCW6FbmNtw': {'datestamp': '2020-04-30 19:08:13.901610', 'hashval': 1119078614239620630,
                          'note': 'Pick up letter'},
 '-M6AN37PRF0p66IT6BeF': {'datestamp': '2020-04-29 19:26:51.509552', 'hashval': -8680445577415429021,
                          'note': 'Go to gym and do yoga'},
 '-M6AN5rGyt35SgFrv8VJ': {'datestamp': '2020-04-29 19:27:02.669504', 'hashval': 3918774690137044647,
                          'note': 'Buy bread.'}}

'''